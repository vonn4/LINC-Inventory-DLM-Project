import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

def apply_sheet_formatting(workbook, sheet_name, header_color, data_color=None):
    """Apply color formatting to Excel sheets"""
    if sheet_name not in workbook.sheetnames:
        return
    
    ws = workbook[sheet_name]
    
    # Define color fills
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply header formatting (first row)
    if ws.max_row > 0:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
    
    # Apply alternating row colors if data_color is provided
    if data_color and ws.max_row > 1:
        data_fill = PatternFill(start_color=data_color, end_color=data_color, fill_type="solid")
        for row_num in range(2, ws.max_row + 1, 2):  # Every other row starting from row 2
            for cell in ws[row_num]:
                cell.fill = data_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

def calculate_device_age_risk(age_years):
    """
    Calculate risk score based on device age (Most Important - 50 points max)
    Based on the flowchart: 5+ yrs = High Risk, 3-5 yrs = Medium Risk, <3 yrs = Low Risk
    """
    if pd.isna(age_years):
        return 0, 'Unknown Age'
    
    if age_years >= 5:
        return 50, 'High Risk (5+ years old)'
    elif age_years >= 3:
        return 25, 'Medium Risk (3-5 years old)'
    else:
        return 5, 'Low Risk (<3 years old)'

def calculate_brand_risk(brand):
    """
    Calculate risk score based on brand reliability/support (Important - 30 points max)
    Tier 1: Enterprise brands with excellent support
    Tier 2: Consumer brands with good support  
    Tier 3: Lesser known or discontinued brands
    """
    if pd.isna(brand) or str(brand).strip() == '':
        return 30, 'High Risk (Unknown Brand)'
    
    brand_str = str(brand).lower().strip()
    
    # Tier 1: Enterprise/Premium brands (Low Risk)
    tier1_brands = ['hp', 'dell', 'lenovo', 'apple', 'microsoft', 'cisco', 'canon', 
                    'fujitsu', 'lg', 'samsung', 'sony', 'xerox', 'epson']
    
    # Tier 2: Reliable consumer brands (Medium Risk)
    tier2_brands = ['acer', 'asus', 'logitech', 'netgear', 'linksys', 'viewsonic', 
                    'optoma', 'western digital', 'wd', 'seagate', 'nikon', 'olympus']
    
    # Check brand tier
    if any(tier1 in brand_str for tier1 in tier1_brands):
        return 5, 'Low Risk (Premium Brand)'
    elif any(tier2 in brand_str for tier2 in tier2_brands):
        return 15, 'Medium Risk (Consumer Brand)'
    else:
        return 25, 'High Risk (Lesser Known Brand)'

def calculate_category_risk(category):
    """
    Calculate risk score based on device category criticality (Moderate - 20 points max)
    Critical: Servers, network equipment, medical devices
    Important: Desktops, laptops, printers, monitors
    Standard: Accessories, mobile devices, misc equipment
    """
    if pd.isna(category) or str(category).strip() == '':
        return 20, 'High Risk (Unknown Category)'
    
    category_str = str(category).lower().strip()
    
    # Critical infrastructure devices (High Risk if old)
    critical_categories = ['server', 'network firewall', 'network router', 'network switch', 
                          'network wap', 'defibrillator', 'ups']
    
    # Important business devices (Medium Risk)
    important_categories = ['desktop', 'laptop', 'printer', 'monitor', 'projector', 
                           'phone ip', 'timeclock']
    
    # Standard/accessory devices (Low Risk)
    standard_categories = ['tablet', 'phone cell', 'phone bluetooth', 'webcam', 'speakers', 
                          'camera', 'camcorder', 'charger', 'computer accessory', 
                          'phone accessory', 'docking station']
    
    if any(critical in category_str for critical in critical_categories):
        return 20, 'High Risk (Critical Infrastructure)'
    elif any(important in category_str for important in important_categories):
        return 10, 'Medium Risk (Business Essential)'
    elif any(standard in category_str for standard in standard_categories):
        return 3, 'Low Risk (Standard Equipment)'
    else:
        return 15, 'Medium Risk (Unclassified Category)'

def analyze_device_lifecycle_risk(input_excel_path, output_excel_path):
    """
    Analyze device lifecycle management risk using the Fully_Valid_Data sheet
    """
    try:
        # Read the Fully_Valid_Data sheet
        df = pd.read_excel(input_excel_path, sheet_name='Fully_Valid_Data')
        print(f"Successfully loaded {len(df)} fully valid devices for DLM risk analysis")
        
        # Print available columns for debugging
        print(f"Available columns: {list(df.columns)}")
        
        if len(df) == 0:
            print("No fully valid devices found. Please run the main analyzer first.")
            return
            
    except Exception as e:
        print(f"Error reading input file: {e}")
        print("Make sure you've run the main device analyzer first to create the input file.")
        return
    
    # Calculate risk scores for each factor
    print("\n=== CALCULATING DEVICE LIFECYCLE RISK SCORES ===")
    
    # Age Risk (50 points max - Most Important)
    age_risk_results = df['Device_Age_Years'].apply(calculate_device_age_risk)
    df['Age_Risk_Score'] = [result[0] for result in age_risk_results]
    df['Age_Risk_Reason'] = [result[1] for result in age_risk_results]
    
    # Brand Risk (30 points max - Important)
    brand_risk_results = df['Normalized Brand'].apply(calculate_brand_risk)
    df['Brand_Risk_Score'] = [result[0] for result in brand_risk_results]
    df['Brand_Risk_Reason'] = [result[1] for result in brand_risk_results]
    
    # Category Risk (20 points max - Moderate)
    category_risk_results = df['Normalized Category'].apply(calculate_category_risk)
    df['Category_Risk_Score'] = [result[0] for result in category_risk_results]
    df['Category_Risk_Reason'] = [result[1] for result in category_risk_results]
    
    # Calculate Total Risk Score (0-100 scale)
    df['Total_Risk_Score'] = df['Age_Risk_Score'] + df['Brand_Risk_Score'] + df['Category_Risk_Score']
    
    # Classify overall risk level
    def classify_risk_level(score):
        if score >= 70:
            return 'HIGH RISK'
        elif score >= 35:
            return 'MEDIUM RISK'
        else:
            return 'LOW RISK'
    
    df['Risk_Level'] = df['Total_Risk_Score'].apply(classify_risk_level)
    
    # Add priority ranking within each risk level (by total score)
    df['Priority_Rank'] = df.groupby('Risk_Level')['Total_Risk_Score'].rank(method='dense', ascending=False).astype(int)
    
    # Separate devices by risk level
    high_risk = df[df['Risk_Level'] == 'HIGH RISK'].sort_values('Total_Risk_Score', ascending=False)
    medium_risk = df[df['Risk_Level'] == 'MEDIUM RISK'].sort_values('Total_Risk_Score', ascending=False)
    low_risk = df[df['Risk_Level'] == 'LOW RISK'].sort_values('Total_Risk_Score', ascending=False)
    
    # Display results
    print(f"\n=== DEVICE LIFECYCLE MANAGEMENT RISK ANALYSIS RESULTS ===")
    print(f"🔴 HIGH RISK devices: {len(high_risk)} ({len(high_risk)/len(df)*100:.1f}%)")
    print(f"🟡 MEDIUM RISK devices: {len(medium_risk)} ({len(medium_risk)/len(df)*100:.1f}%)")
    print(f"🟢 LOW RISK devices: {len(low_risk)} ({len(low_risk)/len(df)*100:.1f}%)")
    
    if len(high_risk) > 0:
        print(f"\n🚨 TOP 5 HIGHEST RISK DEVICES:")
        # Use Asset Tag ID as the device identifier
        identifier_col = 'Asset Tag ID'
        display_cols = [identifier_col, 'Normalized Brand', 'Normalized Category', 
                       'Device_Age_Years', 'Total_Risk_Score', 'Age_Risk_Reason']
        # Only include columns that actually exist
        available_cols = [col for col in display_cols if col in df.columns]
        top_high_risk = high_risk.head(5)[available_cols]
        
        for idx, device in top_high_risk.iterrows():
            device_id = device.get('Asset Tag ID', f"Row {idx}")
            brand = device.get('Normalized Brand', 'Unknown')
            category = device.get('Normalized Category', 'Unknown')
            age = device.get('Device_Age_Years', 0)
            score = device.get('Total_Risk_Score', 0)
            print(f"  Asset {device_id}: {brand} {category} " +
                  f"({age:.1f} yrs) - Score: {score}")
    
    # Create risk summary statistics
    risk_summary = {
        'Risk Level': ['HIGH RISK', 'MEDIUM RISK', 'LOW RISK', 'TOTAL'],
        'Device Count': [len(high_risk), len(medium_risk), len(low_risk), len(df)],
        'Percentage': [
            f"{len(high_risk)/len(df)*100:.1f}%",
            f"{len(medium_risk)/len(df)*100:.1f}%", 
            f"{len(low_risk)/len(df)*100:.1f}%",
            "100.0%"
        ],
        'Avg Risk Score': [
            f"{high_risk['Total_Risk_Score'].mean():.1f}" if len(high_risk) > 0 else "N/A",
            f"{medium_risk['Total_Risk_Score'].mean():.1f}" if len(medium_risk) > 0 else "N/A",
            f"{low_risk['Total_Risk_Score'].mean():.1f}" if len(low_risk) > 0 else "N/A",
            f"{df['Total_Risk_Score'].mean():.1f}"
        ],
        'Replacement Priority': [
            'IMMEDIATE (Next 6 months)',
            'PLANNED (6-18 months)', 
            'SCHEDULED (18+ months)',
            'Various'
        ]
    }
    
    # Create additional analysis data first
    print("\n📊 Creating detailed analysis summaries...")
    
    # Most risky brands analysis
    brand_risk_analysis = df.groupby('Normalized Brand').agg({
        'Total_Risk_Score': ['count', 'mean', 'max'],
        'Device_Age_Years': 'mean',
        'Risk_Level': lambda x: (x == 'HIGH RISK').sum()
    }).round(1)
    brand_risk_analysis.columns = ['Device_Count', 'Avg_Risk_Score', 'Max_Risk_Score', 'Avg_Age_Years', 'High_Risk_Count']
    brand_risk_analysis = brand_risk_analysis.sort_values('Avg_Risk_Score', ascending=False).reset_index()
    
    # Most risky categories analysis
    category_risk_analysis = df.groupby('Normalized Category').agg({
        'Total_Risk_Score': ['count', 'mean', 'max'],
        'Device_Age_Years': 'mean',
        'Risk_Level': lambda x: (x == 'HIGH RISK').sum()
    }).round(1)
    category_risk_analysis.columns = ['Device_Count', 'Avg_Risk_Score', 'Max_Risk_Score', 'Avg_Age_Years', 'High_Risk_Count']
    category_risk_analysis = category_risk_analysis.sort_values('Avg_Risk_Score', ascending=False).reset_index()
    
    # Age distribution analysis
    age_distribution = pd.DataFrame({
        'Age_Range': ['0-2 years', '3-4 years', '5-6 years', '7-9 years', '10+ years'],
        'Device_Count': [
            ((df['Device_Age_Years'] >= 0) & (df['Device_Age_Years'] < 3)).sum(),
            ((df['Device_Age_Years'] >= 3) & (df['Device_Age_Years'] < 5)).sum(),
            ((df['Device_Age_Years'] >= 5) & (df['Device_Age_Years'] < 7)).sum(),
            ((df['Device_Age_Years'] >= 7) & (df['Device_Age_Years'] < 10)).sum(),
            (df['Device_Age_Years'] >= 10).sum()
        ]
    })
    age_distribution['Percentage'] = (age_distribution['Device_Count'] / len(df) * 100).round(1)
    age_distribution['Risk_Assessment'] = ['Low Risk', 'Medium Risk', 'High Risk', 'Very High Risk', 'Critical Risk']

    # Age distribution analysis for high-risk devices
    if len(high_risk) > 0:
        very_old = (high_risk['Device_Age_Years'] >= 7).sum()
        old = ((high_risk['Device_Age_Years'] >= 5) & (high_risk['Device_Age_Years'] < 7)).sum()
        print(f"\n🔍 HIGH RISK DEVICE BREAKDOWN:")
        print(f"  📅 Very Old (7+ years): {very_old} devices")
        print(f"  📅 Old (5-7 years): {old} devices")
        
        # Brand analysis for high-risk devices
        high_risk_brands = high_risk['Normalized Brand'].value_counts().head(5)
        print(f"\n🏷️  TOP BRANDS IN HIGH RISK:")
        for brand, count in high_risk_brands.items():
            print(f"  {brand}: {count} devices")
    
    # Display top risky brands overall
    print(f"\n📈 TOP 5 RISKIEST BRANDS (by average risk score):")
    top_risky_brands = brand_risk_analysis.head(5)
    for idx, brand_data in top_risky_brands.iterrows():
        print(f"  {brand_data['Normalized Brand']}: Avg Risk {brand_data['Avg_Risk_Score']:.1f} " +
              f"({brand_data['Device_Count']} devices, {brand_data['High_Risk_Count']} high-risk)")
    
    # Display top risky categories
    print(f"\n📂 TOP 5 RISKIEST CATEGORIES (by average risk score):")
    top_risky_categories = category_risk_analysis.head(5)
    for idx, cat_data in top_risky_categories.iterrows():
        print(f"  {cat_data['Normalized Category']}: Avg Risk {cat_data['Avg_Risk_Score']:.1f} " +
              f"({cat_data['Device_Count']} devices, {cat_data['High_Risk_Count']} high-risk)")
    
    # Display age distribution insights
    print(f"\n📅 DEVICE AGE DISTRIBUTION:")
    for idx, age_data in age_distribution.iterrows():
        print(f"  {age_data['Age_Range']}: {age_data['Device_Count']} devices " +
              f"({age_data['Percentage']:.1f}%) - {age_data['Risk_Assessment']}")
    
    # Save results to Excel
    try:
        with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
            # 1. Complete Risk Analysis - MOVED TO FIRST POSITION
            df.sort_values('Total_Risk_Score', ascending=False).to_excel(
                writer, sheet_name='Complete_Risk_Analysis', index=False)
            
            # 2. Risk Summary Dashboard
            pd.DataFrame(risk_summary).to_excel(writer, sheet_name='Risk_Summary_Dashboard', index=False)
            
            # 3. Brand Risk Analysis
            brand_risk_analysis.to_excel(writer, sheet_name='Brand_Risk_Analysis', index=False)
            
            # 4. Category Risk Analysis  
            category_risk_analysis.to_excel(writer, sheet_name='Category_Risk_Analysis', index=False)
            
            # 5. Age Distribution Analysis
            age_distribution.to_excel(writer, sheet_name='Age_Distribution_Analysis', index=False)
            
            # 6. High Risk Devices (RED)
            if len(high_risk) > 0:
                high_risk.to_excel(writer, sheet_name='HIGH_RISK_Devices', index=False)
            
            # 7. Medium Risk Devices (YELLOW)  
            if len(medium_risk) > 0:
                medium_risk.to_excel(writer, sheet_name='MEDIUM_RISK_Devices', index=False)
            
            # 8. Low Risk Devices (GREEN)
            if len(low_risk) > 0:
                low_risk.to_excel(writer, sheet_name='LOW_RISK_Devices', index=False)
        
        # Apply color formatting
        print("\n🎨 Applying color formatting to DLM risk analysis...")
        workbook = openpyxl.load_workbook(output_excel_path)
        
        # Color scheme for risk levels and analysis sheets
        sheet_colors = {
            'Complete_Risk_Analysis': ('2C3E50', 'EBF5FB'),      # Dark Blue - Primary Analysis
            'Risk_Summary_Dashboard': ('8E44AD', 'E8DAEF'),      # Purple - Executive Summary
            'Brand_Risk_Analysis': ('D68910', 'FEF9E7'),         # Orange - Brand Analysis
            'Category_Risk_Analysis': ('148F77', 'E8F8F5'),      # Teal - Category Analysis
            'Age_Distribution_Analysis': ('5B2C6F', 'F4ECF7'),   # Deep Purple - Age Analysis
            'HIGH_RISK_Devices': ('C0392B', 'F5B7B1'),          # Bright Red - Critical
            'MEDIUM_RISK_Devices': ('F39C12', 'FCF3CF'),        # Yellow/Orange - Caution  
            'LOW_RISK_Devices': ('27AE60', 'D5F4E6')            # Green - Safe
        }
        
        # Apply formatting to each sheet
        sheets_formatted = 0
        for sheet_name, (header_color, data_color) in sheet_colors.items():
            if sheet_name in workbook.sheetnames:
                apply_sheet_formatting(workbook, sheet_name, header_color, data_color)
                sheets_formatted += 1
                print(f"  ✅ Formatted {sheet_name}")
        
        workbook.save(output_excel_path)
        workbook.close()
        print(f"🎨 Applied color formatting to {sheets_formatted} sheets!")
        
        print(f"\n✅ Device Lifecycle Management risk analysis saved to: {output_excel_path}")
        print(f"\n📊 EXECUTIVE SUMMARY:")
        print(f"   🔴 {len(high_risk)} devices need IMMEDIATE attention (replacement within 6 months)")
        print(f"   🟡 {len(medium_risk)} devices need PLANNED replacement (6-18 months)")  
        print(f"   🟢 {len(low_risk)} devices are in good condition (18+ months)")
        
    except Exception as e:
        print(f"\nError saving DLM risk analysis: {e}")

def main():
    # File paths
    input_file = r'C:\Users\AbrehamMesfin\Downloads\device_analysis_with_categories.xlsx'
    output_file = r'C:\Users\AbrehamMesfin\Downloads\device_lifecycle_risk_analysis.xlsx'
    
    print("🔄 Device Lifecycle Management (DLM) Risk Analyzer")
    print("=" * 60)
    print("This tool analyzes devices from the 'Fully_Valid_Data' sheet and")
    print("creates a risk-based classification for lifecycle management.")
    print("\n📋 Risk Scoring System:")
    print("   🕐 Device Age (50 points max): 5+ years = High, 3-5 years = Medium, <3 years = Low")
    print("   🏷️  Brand Reliability (30 points max): Enterprise > Consumer > Unknown")
    print("   📂 Device Category (20 points max): Critical > Business > Standard")
    print("   📊 Total Risk: 70+ = HIGH, 35-69 = MEDIUM, <35 = LOW")
    
    analyze_device_lifecycle_risk(input_file, output_file)

if __name__ == "__main__":
    main()
