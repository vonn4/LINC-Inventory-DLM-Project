import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

def read_device_data(csv_path):
    return pd.read_csv(csv_path, encoding='latin-1')

def apply_sheet_formatting(workbook, sheet_name, header_color, data_color=None):
    """Apply color formatting to Excel sheets"""
    if sheet_name not in workbook.sheetnames:
        return
    
    ws = workbook[sheet_name]
    
    # Define color fills
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply header formatting (first row)
    if ws.max_row > 0:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
    
    # Apply alternating row colors if data_color is provided
    if data_color and ws.max_row > 1:
        data_fill = PatternFill(start_color=data_color, end_color=data_color, fill_type="solid")
        for row_num in range(2, ws.max_row + 1, 2):  # Every other row starting from row 2
            for cell in ws[row_num]:
                cell.fill = data_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

def main():
    csv_path = r'C:\Users\AbrehamMesfin\Downloads\Inventory.csv'
    output_path = r'C:\Users\AbrehamMesfin\Downloads\device_analysis_with_categories.xlsx'
    
    try:
        df = read_device_data(csv_path)
        print(f"Successfully loaded data with {len(df)} rows")
    except FileNotFoundError:
        print(f"Error: Could not find the CSV file at {csv_path}")
        print("Please make sure the file exists or update the csv_path variable.")
        return
    except Exception as e:
        print(f"Error loading data: {e}")
        return

    # Devices with empty or missing Brand BEFORE normalization
    unrecognized_brands = df[df['Brand'].isna() | (df['Brand'].astype(str).str.strip() == '')]

    def normalize_brand(brand):
        if pd.isna(brand):
            return ""
        brand_str = str(brand).strip().lower()
        # Handle common variations and misspellings
        brand_replacements = {
            'epsson': 'epson',                                                                                        
            'tripplite': 'tripp lite',
            'hewlett packard': 'hp'
        }
        # Apply replacements
        for old, new in brand_replacements.items():
            if brand_str == old:
                brand_str = new
        return brand_str.replace('-', ' ').replace('_', ' ')

    # Extract and normalize all unique brand names from the Brand column
    brand_name = set(df['Brand'].dropna().apply(normalize_brand).unique())
    
    # Add normalized brand column
    df['Normalized Brand'] = df['Brand'].apply(normalize_brand)

    # Remove unrecognized devices from main DataFrame
    recognized_brands = df[~(df['Brand'].isna() | (df['Brand'].astype(str).str.strip() == ''))]
    
    # === NEW CATEGORY NORMALIZATION SECTION ===
    
    # Devices with empty or missing Category BEFORE normalization
    unrecognized_categories = df[df['Category'].isna() | (df['Category'].astype(str).str.strip() == '')]

    def normalize_category(category):
        if pd.isna(category):
            return ""
        category_str = str(category).strip().lower()
        # Handle common variations
        category_replacements = {
            'defibulator': 'defibrillator',  # Fix spelling
            'pc desktop': 'desktop',
            'pc laptop': 'laptop'
        }
        # Apply replacements
        for old, new in category_replacements.items():
            if category_str == old:
                category_str = new
        return category_str.replace('-', ' ').replace('_', ' ')

    # Extract and normalize all unique category names from the Category column
    category_names = set(df['Category'].dropna().apply(normalize_category).unique())
    
    # Add normalized category column
    df['Normalized Category'] = df['Category'].apply(normalize_category)

    # Remove unrecognized categories from main DataFrame
    recognized_categories = df[~(df['Category'].isna() | (df['Category'].astype(str).str.strip() == ''))]
    
    # === END NEW CATEGORY SECTION ===
    
    # === PURCHASE DATE VALIDATION AND AGE ANALYSIS ===
    
    # Parse purchase dates and validate them
    def validate_purchase_date(date_str):
        if pd.isna(date_str):
            return None, 'Missing'
        try:
            parsed_date = pd.to_datetime(date_str)
            current_date = pd.Timestamp.now()
            
            # Check if date is in the future
            if parsed_date > current_date:
                return None, 'Future Date'
            
            # Check if date is unreasonably old (before 2010)
            if parsed_date.year < 2010:
                return None, 'Too Old'
            
            return parsed_date, 'Valid'
            
        except:
            return None, 'Invalid Format'
    
    # Apply validation
    validation_results = df['Purchase Date'].apply(validate_purchase_date)
    df['Purchase_Date_Parsed'] = [result[0] for result in validation_results]
    df['Purchase_Date_Status'] = [result[1] for result in validation_results]
    
    # Calculate age only for valid dates
    current_date = pd.Timestamp.now()
    df['Device_Age_Years'] = (
        (current_date - df['Purchase_Date_Parsed']).dt.days / 365.25
    ).round(1)
    
    # Separate devices based on purchase date validity
    valid_purchase_dates = df[df['Purchase_Date_Status'] == 'Valid']
    invalid_purchase_dates = df[df['Purchase_Date_Status'] != 'Valid']
    
    # === END PURCHASE DATE SECTION ===
    
    # Show brand results
    print("=== BRAND NORMALIZATION RESULTS ===")
    print("Unique normalized brands found in the sheet:")
    for b in sorted(brand_name):
        print(f"  {b}")
    
    print(f"\nTotal unique brands found: {len(brand_name)}")
    print(f"Devices with unrecognized brands: {len(unrecognized_brands)}")
    print(f"Devices with recognized brands: {len(recognized_brands)}")
    
    # Show category results
    print("\n=== CATEGORY NORMALIZATION RESULTS ===")
    print("Unique normalized categories found in the sheet:")
    for c in sorted(category_names):
        print(f"  {c}")
    
    print(f"\nTotal unique categories found: {len(category_names)}")
    print(f"Devices with unrecognized categories: {len(unrecognized_categories)}")
    print(f"Devices with recognized categories: {len(recognized_categories)}")
    
    # Show purchase date and age results
    print("\n=== PURCHASE DATE VALIDATION RESULTS ===")
    
    # Show validation status breakdown
    status_counts = df['Purchase_Date_Status'].value_counts()
    print("Purchase Date Status Breakdown:")
    for status, count in status_counts.items():
        percentage = round((count / len(df) * 100), 1)
        print(f"  {status}: {count} devices ({percentage}%)")
    
    print(f"\nDevices with VALID purchase dates: {len(valid_purchase_dates)} ({round(len(valid_purchase_dates)/len(df)*100, 1)}%)")
    print(f"Devices with INVALID purchase dates: {len(invalid_purchase_dates)} ({round(len(invalid_purchase_dates)/len(df)*100, 1)}%)")
    
    # Analysis for valid purchase dates only
    if len(valid_purchase_dates) > 0:
        print(f"\n=== AGE ANALYSIS (Valid Purchase Dates Only) ===")
        print(f"Average device age: {valid_purchase_dates['Device_Age_Years'].mean():.1f} years")
        print(f"Oldest device: {valid_purchase_dates['Device_Age_Years'].max():.1f} years")
        print(f"Newest device: {valid_purchase_dates['Device_Age_Years'].min():.1f} years")
        
        # Show age distribution for valid dates only
        age_ranges = [
            (0, 1, 'Less than 1 year'),
            (1, 3, '1-3 years'),
            (3, 5, '3-5 years'),
            (5, 10, '5-10 years'),
            (10, float('inf'), 'Over 10 years')
        ]
        
        print(f"\nAge Distribution (Valid Dates Only):")
        for min_age, max_age, label in age_ranges:
            if max_age == float('inf'):
                count = (valid_purchase_dates['Device_Age_Years'] >= min_age).sum()
            else:
                count = ((valid_purchase_dates['Device_Age_Years'] >= min_age) & 
                        (valid_purchase_dates['Device_Age_Years'] < max_age)).sum()
            percentage = round((count / len(valid_purchase_dates) * 100), 1)
            print(f"  {label}: {count} devices ({percentage}%)")
    
    # Show invalid date details
    if len(invalid_purchase_dates) > 0:
        print(f"\n=== INVALID PURCHASE DATE DETAILS ===")
        invalid_status_counts = invalid_purchase_dates['Purchase_Date_Status'].value_counts()
        for status, count in invalid_status_counts.items():
            print(f"  {status}: {count} devices")
            # Show a few examples of invalid dates
            examples = invalid_purchase_dates[invalid_purchase_dates['Purchase_Date_Status'] == status]['Purchase Date'].dropna().head(3).tolist()
            if examples:
                print(f"    Examples: {examples}")
    
    # === DATA QUALITY INSIGHTS ===
    print(f"\n=== DATA QUALITY INSIGHTS & RECOMMENDATIONS ===")
    
    # Overall data quality score
    fully_valid_count = len(df[
        ~(df['Brand'].isna() | (df['Brand'].astype(str).str.strip() == '')) &
        ~(df['Category'].isna() | (df['Category'].astype(str).str.strip() == '')) &
        (df['Purchase_Date_Status'] == 'Valid')
    ])
    data_quality_score = round((fully_valid_count / len(df)) * 100, 1)
    print(f"📊 Overall Data Quality Score: {data_quality_score}% ({fully_valid_count}/{len(df)} devices fully valid)")
    
    # Specific recommendations
    if len(unrecognized_brands) > 0:
        print(f"🏷️  Brand Issues: {len(unrecognized_brands)} devices need brand cleanup")
    if len(unrecognized_categories) > 0:
        print(f"📂 Category Issues: {len(unrecognized_categories)} devices need category assignment")
    if len(invalid_purchase_dates) > 0:
        print(f"📅 Date Issues: {len(invalid_purchase_dates)} devices need purchase date correction")
    
    # Age-based recommendations for lifecycle management
    if len(valid_purchase_dates) > 0:
        old_devices = (valid_purchase_dates['Device_Age_Years'] >= 5).sum()
        very_old_devices = (valid_purchase_dates['Device_Age_Years'] >= 10).sum()
        
        if old_devices > 0:
            print(f"⚠️  Lifecycle Alert: {old_devices} devices are 5+ years old (consider replacement planning)")
        if very_old_devices > 0:
            print(f"🚨 Critical Age Alert: {very_old_devices} devices are 10+ years old (replacement recommended)")
    
    print(f"✅ Ready for DLM Analysis: {fully_valid_count} devices with complete, valid data")
    
    # Save all results to Excel with multiple sheets
    try:
        # First, save data to Excel without formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Original data sheet - unprocessed data for comparison
            df.to_excel(writer, sheet_name='Original_Data', index=False)
            
            # Brand sheets
            recognized_brands.to_excel(writer, sheet_name='All_Brands_Recognized', index=False)
            unrecognized_brands.to_excel(writer, sheet_name='Brands_Unrecognized', index=False)
            
            # Category sheets
            recognized_categories.to_excel(writer, sheet_name='All_Categories_Recognized', index=False)
            unrecognized_categories.to_excel(writer, sheet_name='Categories_Unrecognized', index=False)
            
            # === PURCHASE DATE FOCUSED SHEETS ===
            
            # Valid purchase dates - with age analysis
            if len(valid_purchase_dates) > 0:
                valid_purchase_dates.to_excel(writer, sheet_name='Valid_Purchase_Dates', index=False)
            
            # Invalid purchase dates - separate sheet for review
            if len(invalid_purchase_dates) > 0:
                invalid_purchase_dates.to_excel(writer, sheet_name='Invalid_Purchase_Dates', index=False)
            
            # Combined analysis - devices with valid brand, category AND purchase date
            fully_valid = df[
                ~(df['Brand'].isna() | (df['Brand'].astype(str).str.strip() == '')) &
                ~(df['Category'].isna() | (df['Category'].astype(str).str.strip() == '')) &
                (df['Purchase_Date_Status'] == 'Valid')
            ]
            if len(fully_valid) > 0:
                fully_valid.to_excel(writer, sheet_name='Fully_Valid_Data', index=False)
            
            # All invalid data - devices with ANY invalid data (brand, category, or purchase date)
            all_invalid = df[
                (df['Brand'].isna() | (df['Brand'].astype(str).str.strip() == '')) |
                (df['Category'].isna() | (df['Category'].astype(str).str.strip() == '')) |
                (df['Purchase_Date_Status'] != 'Valid')
            ]
            if len(all_invalid) > 0:
                # Add a column showing what issues each device has
                def identify_issues(row):
                    issues = []
                    if pd.isna(row['Brand']) or str(row['Brand']).strip() == '':
                        issues.append('Missing Brand')
                    if pd.isna(row['Category']) or str(row['Category']).strip() == '':
                        issues.append('Missing Category')
                    if row['Purchase_Date_Status'] != 'Valid':
                        issues.append(f'Invalid Purchase Date ({row["Purchase_Date_Status"]})')
                    return ' | '.join(issues)
                
                all_invalid = all_invalid.copy()
                all_invalid['Issues_Found'] = all_invalid.apply(identify_issues, axis=1)
                all_invalid.to_excel(writer, sheet_name='All_Invalid_Data', index=False)
            
            # Overall Data Quality Summary - comprehensive overview
            summary_data = {
                'Data Category': ['Brands', 'Categories', 'Purchase Dates', 'Overall Data Quality'],
                'Valid Count': [
                    len(recognized_brands),                                                                                               
                    len(recognized_categories), 
                    len(valid_purchase_dates),
                    len(fully_valid)
                ],
                'Invalid Count': [
                    len(unrecognized_brands),
                    len(unrecognized_categories),
                    len(invalid_purchase_dates),
                    len(df) - len(fully_valid)
                ],
                'Total Devices': [
                    len(df),
                    len(df),
                    len(df),
                    len(df)                                     
                ],
                'Valid Percentage': [
                    round((len(recognized_brands) / len(df)) * 100, 1),                                 
                    round((len(recognized_categories) / len(df)) * 100, 1),                                                                                                                                                                                                                         
                    round((len(valid_purchase_dates) / len(df)) * 100, 1),
                    round((len(fully_valid) / len(df)) * 100, 1)
                ],
                'Invalid Percentage': [
                    round((len(unrecognized_brands) / len(df)) * 100, 1),
                    round((len(unrecognized_categories) / len(df)) * 100, 1),
                    round((len(invalid_purchase_dates) / len(df)) * 100, 1),
                    round(((len(df) - len(fully_valid)) / len(df)) * 100, 1)                                                                       
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Data_Quality_Summary', index=False)
        
        # Now apply color formatting
        print("📝 Applying color  formatting to Excel sheets...")
        workbook = openpyxl.load_workbook(output_path)
        
        # Color scheme for different sheet types
        sheet_colors = {
            'Original_Data': ('366092', 'D9E2F3'),          # Blue theme - original data
            'All_Brands_Recognized': ('70AD47', 'E2EFDA'),  # Green theme - valid data
            'Brands_Unrecognized': ('E74C3C', 'FADBD8'),    # Red theme - invalid data
            'All_Categories_Recognized': ('70AD47', 'E2EFDA'),  # Green theme - valid data
            'Categories_Unrecognized': ('E74C3C', 'FADBD8'), # Red theme - invalid data
            'Valid_Purchase_Dates': ('70AD47', 'E2EFDA'),   # Green theme - valid data
            'Invalid_Purchase_Dates': ('E74C3C', 'FADBD8'), # Red theme - invalid data
            'Fully_Valid_Data': ('27AE60', 'D5F4E6'),       # Bright green - best data
            'All_Invalid_Data': ('C0392B', 'F5B7B1'),       # Bright red - problem data
            'Data_Quality_Summary': ('8E44AD', 'E8DAEF')    # Purple theme - summary/analysis
        }
        
        # Apply formatting to each sheet
        sheets_formatted = 0
        for sheet_name, (header_color, data_color) in sheet_colors.items():
            if sheet_name in workbook.sheetnames:
                apply_sheet_formatting(workbook, sheet_name, header_color, data_color)
                sheets_formatted += 1
                print(f"  ✅ Formatted {sheet_name}")
        
        # Save the formatted workbook
        workbook.save(output_path)
        workbook.close()
        print(f"🎨 Applied color formatting to {sheets_formatted} sheets!")

        print(f"\nResults saved to {output_path}")
        print(f"Devices with valid purchase dates: {len(valid_purchase_dates)}")
        print(f"Devices with invalid purchase dates: {len(invalid_purchase_dates)}")
        if len(valid_purchase_dates) > 0 and len(invalid_purchase_dates) > 0:
            print(f"Devices with brand, category AND valid purchase date: {len(fully_valid) if 'fully_valid' in locals() else 0}")
        if 'all_invalid' in locals() and len(all_invalid) > 0:
            print(f"Devices with ANY invalid data (brand, category, or purchase date): {len(all_invalid)}")
        
    except PermissionError:
        print(f"\nERROR: Permission denied when trying to save to {output_path}")
        print("This usually means:")
        print("1. The file is currently open in Excel or another application - please close it")
        print("2. You don't have write permissions to the folder")
        print("3. The file is locked by another process")
        print("\nPlease close any open Excel files and try again.")
        
    except Exception as e:
        print(f"\nError saving results: {e}")

if __name__ == "__main__":
    main()
