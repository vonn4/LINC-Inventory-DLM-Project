import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

def read_device_data(csv_path):
    return pd.read_csv(csv_path, encoding='latin-1')

def apply_sheet_formatting(workbook, sheet_name, header_color, data_color=None):
    """Apply color formatting to Excel sheets with improved error handling"""
    try:
        if sheet_name not in workbook.sheetnames:
            print(f"  ⚠️  Sheet '{sheet_name}' not found in workbook")
            return False
        
        ws = workbook[sheet_name]
        print(f"  🎨 Formatting sheet '{sheet_name}' with {ws.max_row} rows and {ws.max_column} columns")
        
        # Ensure we have valid hex colors (remove # if present and ensure 6 digits)
        if header_color.startswith('#'):
            header_color = header_color[1:]
        if len(header_color) != 6:
            print(f"  ⚠️  Invalid header color: {header_color}, using default")
            header_color = "366092"
            
        if data_color:
            if data_color.startswith('#'):
                data_color = data_color[1:]
            if len(data_color) != 6:
                print(f"  ⚠️  Invalid data color: {data_color}, using default")
                data_color = "D9E2F3"
        
        # Define color fills
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply header formatting (first row)
        if ws.max_row > 0:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
        
        # Apply alternating row colors if data_color is provided
        if data_color and ws.max_row > 1:
            data_fill = PatternFill(start_color=data_color, end_color=data_color, fill_type="solid")
            for row_num in range(2, ws.max_row + 1, 2):  # Every other row starting from row 2
                for cell in ws[row_num]:
                    cell.fill = data_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return True
        
    except Exception as e:
        print(f"  ❌ Error formatting sheet '{sheet_name}': {e}")
        return False

def main():
    csv_path = 'Inventory.csv'  # Use relative path since it's in the same directory
    output_path = r'C:\Users\AbrehamMesfin\OneDrive - Greater KC LINC, Inc\Documents\VS code API project\device_analysis_with_categories.xlsx'
    
    try:
        df = read_device_data(csv_path)
        print(f"Successfully loaded data with {len(df)} rows")
        
        # Keep a copy of the original data completely unchanged
        original_df = df.copy()
        
    except FileNotFoundError:
        print(f"Error: Could not find the CSV file at {csv_path}")
        print("Please make sure the file exists or update the csv_path variable.")
        return
    except Exception as e:
        print(f"Error loading data: {e}")
        return

    # === DEVICE STATUS VALIDATION (NEW - BEFORE BRAND PROCESSING) ===
    
    # Normalize and categorize device status
    def normalize_status(status):
        if pd.isna(status):
            return "Unknown"
        status_str = str(status).strip().lower()
        
        # Active/Available statuses (Green - devices in use or ready for use)
        active_statuses = [
            'available', 'check out', 'checked out', 'check in', 'checked in',
            'under repair', 'found', 'reserved'
        ]
        
        # Inactive/Unavailable statuses (Red - devices no longer in active inventory)
        inactive_statuses = [
            'broken', 'lost/missing', 'lost', 'missing', 'donate', 'donated',
            'dispose', 'disposed', 'sold'
        ]
        
        # Check for active statuses
        for active in active_statuses:
            if active in status_str:
                return f"ACTIVE ({status.strip()})"
        
        # Check for inactive statuses  
        for inactive in inactive_statuses:
            if inactive in status_str:
                return f"INACTIVE ({status.strip()})"
        
        # Unknown status
        return f"UNKNOWN ({status.strip()})"
    
    # Apply status normalization
    df['Status_Normalized'] = df['Status'].apply(normalize_status)
    
    # Separate devices by status availability
    available_active_devices = df[df['Status_Normalized'].str.startswith('ACTIVE')]
    unavailable_inactive_devices = df[df['Status_Normalized'].str.startswith('INACTIVE')] 
    unknown_status_devices = df[df['Status_Normalized'].str.startswith('UNKNOWN')]
    
    print("=== DEVICE STATUS AVAILABILITY RESULTS ===")
    print(f"✅ AVAILABLE/ACTIVE devices: {len(available_active_devices)} ({len(available_active_devices)/len(df)*100:.1f}%)")
    print(f"❌ UNAVAILABLE/INACTIVE devices: {len(unavailable_inactive_devices)} ({len(unavailable_inactive_devices)/len(df)*100:.1f}%)")
    print(f"❓ UNKNOWN STATUS devices: {len(unknown_status_devices)} ({len(unknown_status_devices)/len(df)*100:.1f}%)")
    
    # Show status breakdown
    status_counts = df['Status_Normalized'].value_counts()
    print("\nDetailed Status Breakdown:")
    for status, count in status_counts.items():
        percentage = round((count / len(df) * 100), 1)
        print(f"  {status}: {count} devices ({percentage}%)")
    
    # === END DEVICE STATUS SECTION ===
    
    # Devices with empty or missing Brand BEFORE normalization
    unrecognized_brands = df[df['Brand'].isna() | (df['Brand'].astype(str).str.strip() == '')]

    def normalize_brand(brand):
        if pd.isna(brand):
            return ""
        brand_str = str(brand).strip().lower()
        # Handle common variations and misspellings
        brand_replacements = {
            'epsson': 'epson',                                                                                        
            'tripplite': 'tripp lite',
            'hewlett packard': 'hp'
        }
        # Apply replacements
        for old, new in brand_replacements.items():
            if brand_str == old:
                brand_str = new
        # Clean and format with proper capitalization
        cleaned_brand = brand_str.replace('-', ' ').replace('_', ' ')
        return cleaned_brand.title() if cleaned_brand else ""

    # Extract and normalize all unique brand names from the Brand column
    brand_name = set(df['Brand'].dropna().apply(normalize_brand).unique())
    
    # Update the existing Brand column with normalized values
    df['Brand'] = df['Brand'].apply(normalize_brand)

    # Remove unrecognized devices from main DataFrame
    recognized_brands = df[~(df['Brand'].isna() | (df['Brand'].astype(str).str.strip() == ''))]
    
    # === NEW CATEGORY NORMALIZATION SECTION ===
    
    # Devices with empty or missing Category BEFORE normalization
    unrecognized_categories = df[df['Category'].isna() | (df['Category'].astype(str).str.strip() == '')]

    def normalize_category(category):
        if pd.isna(category):
            return ""
        category_str = str(category).strip().lower()
        # Handle common variations
        category_replacements = {
            'defibulator': 'defibrillator',  # Fix spelling
            'pc desktop': 'desktop',
            'pc laptop': 'laptop'
        }
        # Apply replacements
        for old, new in category_replacements.items():
            if category_str == old:
                category_str = new
        # Clean and format with proper capitalization
        cleaned_category = category_str.replace('-', ' ').replace('_', ' ')
        return cleaned_category.title() if cleaned_category else ""

    # Extract and normalize all unique category names from the Category column
    category_names = set(df['Category'].dropna().apply(normalize_category).unique())
    
    # Update the existing Category column with normalized values
    df['Category'] = df['Category'].apply(normalize_category)

    # Remove unrecognized categories from main DataFrame
    recognized_categories = df[~(df['Category'].isna() | (df['Category'].astype(str).str.strip() == ''))]
    
    # === END NEW CATEGORY SECTION ===
    
    # === PURCHASE DATE VALIDATION AND AGE ANALYSIS ===
    
    # Parse purchase dates and validate them
    def validate_purchase_date(date_str):
        if pd.isna(date_str):
            return None, 'Missing'
        try:
            parsed_date = pd.to_datetime(date_str)
            current_date = pd.Timestamp.now()
            
            # Check if date is in the future
            if parsed_date > current_date:
                return None, 'Future Date'
            
            # Check if date is unreasonably old (before 2010)
            if parsed_date.year < 2010:
                return None, 'Too Old'
            
            return parsed_date, 'Valid'
            
        except:
            return None, 'Invalid Format'
    
    # Apply validation
    validation_results = df['Purchase Date'].apply(validate_purchase_date)
    df['Purchase_Date_Parsed'] = [result[0] for result in validation_results]
    df['Purchase_Date_Status'] = [result[1] for result in validation_results]
    
    # Calculate age only for valid dates
    current_date = pd.Timestamp.now()
    df['Device_Age_Years'] = (
        (current_date - df['Purchase_Date_Parsed']).dt.days / 365.25
    ).round(1)
    
    # Separate devices based on purchase date validity
    valid_purchase_dates = df[df['Purchase_Date_Status'] == 'Valid']
    invalid_purchase_dates = df[df['Purchase_Date_Status'] != 'Valid']
    
    # === END PURCHASE DATE SECTION ===
    
    # Show brand results
    print("=== BRAND NORMALIZATION RESULTS ===")
    print("Unique normalized brands found in the sheet:")
    for b in sorted(brand_name):
        print(f"  {b}")
    
    print(f"\nTotal unique brands found: {len(brand_name)}")
    print(f"Devices with unrecognized brands: {len(unrecognized_brands)}")
    print(f"Devices with recognized brands: {len(recognized_brands)}")
    
    # Show category results
    print("\n=== CATEGORY NORMALIZATION RESULTS ===")
    print("Unique normalized categories found in the sheet:")
    for c in sorted(category_names):
        print(f"  {c}")
    
    print(f"\nTotal unique categories found: {len(category_names)}")
    print(f"Devices with unrecognized categories: {len(unrecognized_categories)}")
    print(f"Devices with recognized categories: {len(recognized_categories)}")
    
    # Show purchase date and age results
    print("\n=== PURCHASE DATE VALIDATION RESULTS ===")
    
    # Show validation status breakdown
    status_counts = df['Purchase_Date_Status'].value_counts()
    print("Purchase Date Status Breakdown:")
    for status, count in status_counts.items():
        percentage = round((count / len(df) * 100), 1)
        print(f"  {status}: {count} devices ({percentage}%)")
    
    print(f"\nDevices with VALID purchase dates: {len(valid_purchase_dates)} ({round(len(valid_purchase_dates)/len(df)*100, 1)}%)")
    print(f"Devices with INVALID purchase dates: {len(invalid_purchase_dates)} ({round(len(invalid_purchase_dates)/len(df)*100, 1)}%)")
    
    # Analysis for valid purchase dates only
    if len(valid_purchase_dates) > 0:
        print(f"\n=== AGE ANALYSIS (Valid Purchase Dates Only) ===")
        print(f"Average device age: {valid_purchase_dates['Device_Age_Years'].mean():.1f} years")
        print(f"Oldest device: {valid_purchase_dates['Device_Age_Years'].max():.1f} years")
        print(f"Newest device: {valid_purchase_dates['Device_Age_Years'].min():.1f} years")
        
        # Show age distribution for valid dates only
        age_ranges = [
            (0, 1, 'Less than 1 year'),
            (1, 3, '1-3 years'),
            (3, 5, '3-5 years'),
            (5, 10, '5-10 years'),
            (10, float('inf'), 'Over 10 years')
        ]
        
        print(f"\nAge Distribution (Valid Dates Only):")
        for min_age, max_age, label in age_ranges:
            if max_age == float('inf'):
                count = (valid_purchase_dates['Device_Age_Years'] >= min_age).sum()
            else:
                count = ((valid_purchase_dates['Device_Age_Years'] >= min_age) & 
                        (valid_purchase_dates['Device_Age_Years'] < max_age)).sum()
            percentage = round((count / len(valid_purchase_dates) * 100), 1)
            print(f"  {label}: {count} devices ({percentage}%)")
    
    # Show invalid date details
    if len(invalid_purchase_dates) > 0:
        print(f"\n=== INVALID PURCHASE DATE DETAILS ===")
        invalid_status_counts = invalid_purchase_dates['Purchase_Date_Status'].value_counts()
        for status, count in invalid_status_counts.items():
            print(f"  {status}: {count} devices")
            # Show a few examples of invalid dates
            examples = invalid_purchase_dates[invalid_purchase_dates['Purchase_Date_Status'] == status]['Purchase Date'].dropna().head(3).tolist()
            if examples:
                print(f"    Examples: {examples}")
    
    # === INITIALIZE ENHANCED FULLY VALID DATA ===
    # Create the ORIGINAL fully valid dataset (brand + category + valid purchase date)
    original_fully_valid = df[
        ~(df['Brand'].isna() | (df['Brand'].astype(str).str.strip() == '')) &
        ~(df['Category'].isna() | (df['Category'].astype(str).str.strip() == '')) &
        (df['Purchase_Date_Status'] == 'Valid')
        # NOTE: NOT filtering by status here - that's for final analysis only
    ]
    
    # Start enhanced_fully_valid with the original fully valid data
    enhanced_fully_valid = original_fully_valid.drop(columns=['Purchase_Date_Parsed', 'Purchase_Date_Status', 'Device_Age_Years', 'Status_Normalized'], errors='ignore')
    
    print(f"\n📊 BASELINE FULLY VALID DATA:")
    print(f"   Original fully valid devices (before advanced cleaning): {len(enhanced_fully_valid)}")
    
    # === DATA QUALITY INSIGHTS ===
    print(f"\n=== DATA QUALITY INSIGHTS & RECOMMENDATIONS ===")
    
    # Create analysis-ready data (enhanced_fully_valid + active status filter)
    analysis_ready_devices = enhanced_fully_valid.merge(
        df[['Asset Tag ID', 'Status_Normalized']], 
        on='Asset Tag ID', 
        how='left'
    )
    analysis_ready_devices = analysis_ready_devices[
        analysis_ready_devices['Status_Normalized'].str.startswith('ACTIVE', na=False)
    ].drop(columns=['Status_Normalized'], errors='ignore')
    
    # Overall data quality score - Using analysis-ready devices for final score
    fully_valid_count = len(enhanced_fully_valid)
    analysis_ready_count = len(analysis_ready_devices)
    data_quality_score = round((fully_valid_count / len(df)) * 100, 1)
    analysis_ready_score = round((analysis_ready_count / len(df)) * 100, 1)
    
    print(f"📊 Base Data Quality Score: {data_quality_score}% ({fully_valid_count}/{len(df)} devices fully valid)")
    print(f"🎯 Analysis-Ready Score: {analysis_ready_score}% ({analysis_ready_count}/{len(df)} devices active & fully valid)")
    
    # Specific recommendations
    if len(unrecognized_brands) > 0:
        print(f"🏷️  Brand Issues: {len(unrecognized_brands)} devices need brand cleanup")
    if len(unrecognized_categories) > 0:
        print(f"📂 Category Issues: {len(unrecognized_categories)} devices need category assignment")
    if len(invalid_purchase_dates) > 0:
        print(f"📅 Date Issues: {len(invalid_purchase_dates)} devices need purchase date correction")
    if len(unavailable_inactive_devices) > 0:
        print(f"📱 Status Issues: {len(unavailable_inactive_devices)} devices are inactive/unavailable")
    if len(unknown_status_devices) > 0:
        print(f"❓ Status Unknown: {len(unknown_status_devices)} devices have unknown status")
    
    # Age-based recommendations for lifecycle management
    if len(valid_purchase_dates) > 0:
        old_devices = (valid_purchase_dates['Device_Age_Years'] >= 5).sum()
        very_old_devices = (valid_purchase_dates['Device_Age_Years'] >= 10).sum()
        
        if old_devices > 0:
            print(f"⚠️  Lifecycle Alert: {old_devices} devices are 5+ years old (consider replacement planning)")
        if very_old_devices > 0:
            print(f"🚨 Critical Age Alert: {very_old_devices} devices are 10+ years old (replacement recommended)")
    
    print(f"✅ Ready for DLM Analysis: {analysis_ready_count} devices with complete, valid data and active status")
    
    # Save all results to Excel with multiple sheets
    try:
        # First, save data to Excel without formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Original data sheet - PURE UNMODIFIED inventory data as uploaded
            original_df.to_excel(writer, sheet_name='Original_Data', index=False)
            
            # Brand sheets
            recognized_brands.to_excel(writer, sheet_name='All_Brands_Recognized', index=False)
            unrecognized_brands.to_excel(writer, sheet_name='Brands_Unrecognized', index=False)
            
            # Category sheets
            recognized_categories.to_excel(writer, sheet_name='All_Categories_Recognized', index=False)
            unrecognized_categories.to_excel(writer, sheet_name='Categories_Unrecognized', index=False)
            
            # === NEW STATUS SHEETS ===
            
            # Active/Available devices (Green)
            if len(available_active_devices) > 0:
                available_active_devices.to_excel(writer, sheet_name='Available_Active_Devices', index=False)
            
            # Inactive/Unavailable devices (Red)  
            if len(unavailable_inactive_devices) > 0:
                unavailable_inactive_devices.to_excel(writer, sheet_name='Unavailable_Inactive_Devices', index=False)
            
            # Unknown status devices
            if len(unknown_status_devices) > 0:
                unknown_status_devices.to_excel(writer, sheet_name='Unknown_Status_Devices', index=False)
            
            # === PURCHASE DATE FOCUSED SHEETS ===
            
            # Valid purchase dates - with age analysis
            if len(valid_purchase_dates) > 0:
                valid_purchase_dates.to_excel(writer, sheet_name='Valid_Purchase_Dates', index=False)
            
            # Invalid purchase dates - separate sheet for review
            if len(invalid_purchase_dates) > 0:
                invalid_purchase_dates.to_excel(writer, sheet_name='Invalid_Purchase_Dates', index=False)
            
            
            # Use the enhanced fully_valid data (includes original + corrected devices)
            if len(enhanced_fully_valid) > 0:
                enhanced_fully_valid.to_excel(writer, sheet_name='Fully_Valid_Data', index=False)
            
            # All invalid data - devices with ANY invalid data (brand, category, purchase date, or inactive status)
            all_invalid = df[
                (df['Brand'].isna() | (df['Brand'].astype(str).str.strip() == '')) |
                (df['Category'].isna() | (df['Category'].astype(str).str.strip() == '')) |
                (df['Purchase_Date_Status'] != 'Valid') |
                (~df['Status_Normalized'].str.startswith('ACTIVE'))  # NEW: Include inactive devices
            ]
            if len(all_invalid) > 0:
                # Add a column showing what issues each device has
                def identify_issues(row):
                    issues = []
                    if pd.isna(row['Brand']) or str(row['Brand']).strip() == '':
                        issues.append('Missing Brand')
                    if pd.isna(row['Category']) or str(row['Category']).strip() == '':
                        issues.append('Missing Category')
                    if row['Purchase_Date_Status'] != 'Valid':
                        issues.append(f'Invalid Purchase Date ({row["Purchase_Date_Status"]})')
                    if not row['Status_Normalized'].startswith('ACTIVE'):
                        issues.append(f'Inactive Status ({row["Status_Normalized"]})')
                    return ' | '.join(issues)
                
                all_invalid = all_invalid.copy()
                all_invalid['Issues_Found'] = all_invalid.apply(identify_issues, axis=1)
                all_invalid.to_excel(writer, sheet_name='All_Invalid_Data', index=False)
            
            # Overall Data Quality Summary - comprehensive overview including status
            summary_data = {
                'Data Category': ['Brands', 'Categories', 'Purchase Dates', 'Device Status', 'Fully Valid Data', 'Analysis Ready Data'],
                'Valid Count': [
                    len(recognized_brands),                                                                                               
                    len(recognized_categories), 
                    len(valid_purchase_dates),
                    len(available_active_devices),
                    len(enhanced_fully_valid),
                    len(analysis_ready_devices)
                ],
                'Invalid Count': [
                    len(unrecognized_brands),
                    len(unrecognized_categories),
                    len(invalid_purchase_dates),
                    len(unavailable_inactive_devices) + len(unknown_status_devices),
                    len(df) - len(enhanced_fully_valid),
                    len(df) - len(analysis_ready_devices)
                ],
                'Total Devices': [
                    len(df),
                    len(df),
                    len(df),
                    len(df),
                    len(df),
                    len(df)                                     
                ],
                'Valid Percentage': [
                    round((len(recognized_brands) / len(df)) * 100, 1),                                 
                    round((len(recognized_categories) / len(df)) * 100, 1),                                                                                                                                                                                                                         
                    round((len(valid_purchase_dates) / len(df)) * 100, 1),
                    round((len(available_active_devices) / len(df)) * 100, 1),
                    round((len(enhanced_fully_valid) / len(df)) * 100, 1),
                    round((len(analysis_ready_devices) / len(df)) * 100, 1)
                ],
                'Invalid Percentage': [
                    round((len(unrecognized_brands) / len(df)) * 100, 1),
                    round((len(unrecognized_categories) / len(df)) * 100, 1),
                    round((len(invalid_purchase_dates) / len(df)) * 100, 1),
                    round(((len(unavailable_inactive_devices) + len(unknown_status_devices)) / len(df)) * 100, 1),
                    round(((len(df) - len(enhanced_fully_valid)) / len(df)) * 100, 1),
                    round(((len(df) - len(analysis_ready_devices)) / len(df)) * 100, 1)                                                                      
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Data_Quality_Summary', index=False)
            
         # Analysis-ready data (fully valid + active status) - for DLM risk analysis
            if len(analysis_ready_devices) > 0:
                analysis_ready_devices.to_excel(writer, sheet_name='Analysis_Ready_Data', index=False) 

        # Now apply color formatting
        print("\n🎨 Applying color formatting to Excel sheets...")
        
        try:
            workbook = openpyxl.load_workbook(output_path)
            print(f"📂 Loaded workbook with sheets: {workbook.sheetnames}")
            
            # Color scheme for different sheet types
            sheet_colors = {
                'Original_Data': ('366092', 'D9E2F3'),          # Blue theme - original data
                'All_Brands_Recognized': ('70AD47', 'E2EFDA'),  # Green theme - valid data
                'Brands_Unrecognized': ('E74C3C', 'FADBD8'),    # Red theme - invalid data
                'All_Categories_Recognized': ('70AD47', 'E2EFDA'),  # Green theme - valid data
                'Categories_Unrecognized': ('E74C3C', 'FADBD8'), # Red theme - invalid data
                'Available_Active_Devices': ('27AE60', 'D5F4E6'),    # Bright green - active devices
                'Unavailable_Inactive_Devices': ('E74C3C', 'FADBD8'), # Red theme - inactive devices  
                'Unknown_Status_Devices': ('F39C12', 'FCF3CF'),  # Orange theme - unknown status
                'Valid_Purchase_Dates': ('70AD47', 'E2EFDA'),   # Green theme - valid data
                'Invalid_Purchase_Dates': ('E74C3C', 'FADBD8'), # Red theme - invalid data
                'Fully_Valid_Data': ('27AE60', 'D5F4E6'),       # Bright green - best data
                'Analysis_Ready_Data': ('1F4E79', 'D6EAF8'),    # Deep blue - analysis ready
                'All_Invalid_Data': ('C0392B', 'F5B7B1'),       # Bright red - problem data
                'Enhanced_Fully_Valid_Data': ('27AE60', 'D5F4E6'),  # Bright green - enhanced valid data
                'Remaining_Invalid_Data': ('C0392B', 'F5B7B1'),     # Bright red - remaining invalid data
                'Data_Quality_Summary': ('8E44AD', 'E8DAEF')    # Purple theme - summary/analysis
            }
            
            # Apply formatting to each sheet
            sheets_formatted = 0
            for sheet_name, (header_color, data_color) in sheet_colors.items():
                if sheet_name in workbook.sheetnames:
                    success = apply_sheet_formatting(workbook, sheet_name, header_color, data_color)
                    if success:
                        sheets_formatted += 1
                        print(f"  ✅ Formatted {sheet_name}")
                    else:
                        print(f"  ❌ Failed to format {sheet_name}")
                else:
                    print(f"  ⚠️  Sheet {sheet_name} not found")
            
            # Save the formatted workbook
            workbook.save(output_path)
            workbook.close()
            print(f"🎨 Applied color formatting to {sheets_formatted} sheets!")
            print(f"💾 Saved formatted workbook to: {output_path}")
            
        except Exception as e:
            print(f"❌ Error during color formatting: {e}")
            print("📄 Excel file was saved without color formatting")

        print(f"\nResults saved to {output_path}")
        print(f"Devices with valid purchase dates: {len(valid_purchase_dates)}")
        print(f"Devices with invalid purchase dates: {len(invalid_purchase_dates)}")
        print(f"Devices available/active: {len(available_active_devices)}")
        print(f"Devices unavailable/inactive: {len(unavailable_inactive_devices)}")
        print(f"Enhanced fully valid devices (all corrected): {len(enhanced_fully_valid)}")
        print(f"Analysis-ready devices (active + fully valid): {len(analysis_ready_devices)}")
        if 'remaining_all_invalid' in locals() and len(remaining_all_invalid) > 0:
            print(f"Devices with ANY invalid data (after advanced cleaning): {len(remaining_all_invalid)}")
        print(f"🎯 Ready for DLM Risk Analysis: {len(analysis_ready_devices)} devices")
        
    except PermissionError:
        print(f"\nERROR: Permission denied when trying to save to {output_path}")
        print("This usually means:")
        print("1. The file is currently open in Excel or another application - please close it")
        print("2. You don't have write permissions to the folder")
        print("3. The file is locked by another process")
        print("\nPlease close any open Excel files and try again.")
        
    except Exception as e:
        print(f"\nError saving results: {e}")

    # === ADVANCED DATA CLEANING AND RECLASSIFICATION PROCESS ===
    print(f"\n🔧 === PHASE 1: ADVANCED DATA CLEANING & RECLASSIFICATION ===")
    
    # Enhanced_fully_valid is already initialized above - use it as the starting point
    
    def extract_brand_from_text(text):
        """Extract brand from text using keyword matching"""
        if pd.isna(text):
            return ""
        
        text_lower = str(text).lower().strip()
        
        # Common brand patterns and keywords
        brand_patterns = {
            'apple': ['apple', 'ipad', 'iphone', 'macbook', 'imac'],
            'hp': ['hp', 'hewlett packard', 'pavilion', 'elitebook', 'probook'],
            'dell': ['dell', 'latitude', 'optiplex', 'inspiron', 'precision'],
            'lenovo': ['lenovo', 'thinkpad', 'ideapad', 'yoga'],
            'microsoft': ['microsoft', 'surface', 'xbox'],
            'samsung': ['samsung', 'galaxy'],
            'lg': ['lg electronics', 'lg'],
            'canon': ['canon', 'pixma', 'imageclass'],
            'epson': ['epson', 'workforce', 'expression'],
            'cisco': ['cisco', 'catalyst', 'meraki'],
            'acer': ['acer', 'aspire', 'predator'],
            'asus': ['asus', 'zenbook', 'vivobook'],
            'logitech': ['logitech', 'mx master', 'k400'],
            'sony': ['sony', 'vaio', 'playstation']
        }
        
        for brand, keywords in brand_patterns.items():
            if any(keyword in text_lower for keyword in keywords):
                return brand.title()
        
        return ""
    
    def extract_category_from_text(text):
        """Extract category from text using keyword matching"""
        if pd.isna(text):
            return ""
        
        text_lower = str(text).lower().strip()
        
        # Category patterns and keywords
        category_patterns = {
            'laptop': ['laptop', 'notebook', 'macbook', 'thinkpad', 'elitebook', 'latitude'],
            'desktop': ['desktop', 'pc', 'optiplex', 'imac', 'all-in-one'],
            'tablet': ['tablet', 'ipad', 'surface tablet'],
            'monitor': ['monitor', 'display', 'lcd', 'led monitor'],
            'printer': ['printer', 'pixma', 'laserjet', 'inkjet', 'imageclass'],
            'projector': ['projector', 'beamer'],
            'phone ip': ['ip phone', 'voip', 'desk phone'],
            'phone cell': ['cell phone', 'mobile phone', 'smartphone', 'iphone', 'galaxy'],
            'server': ['server', 'rack server', 'blade server'],
            'network switch': ['switch', 'network switch', 'ethernet switch'],
            'network router': ['router', 'wireless router'],
            'webcam': ['webcam', 'camera', 'web camera'],
            'speakers': ['speakers', 'speaker system', 'audio'],
            'ups': ['ups', 'uninterruptible power', 'battery backup']
        }
        
        for category, keywords in category_patterns.items():
            if any(keyword in text_lower for keyword in keywords):
                return category.title()
        
        return ""
    
    def attempt_data_recovery(device_row):
        """Attempt to recover missing brand/category from other fields"""
        recovered_brand = device_row['Brand'] if not pd.isna(device_row['Brand']) and str(device_row['Brand']).strip() else ""
        recovered_category = device_row['Category'] if not pd.isna(device_row['Category']) and str(device_row['Category']).strip() else ""
        
        # Priority order for data extraction
        extraction_fields = ['Description', 'Device Name', 'Model', 'OS', 'CPU']
        
        # Try to recover brand
        if not recovered_brand:
            for field in extraction_fields:
                if field in device_row and not pd.isna(device_row[field]):
                    extracted_brand = extract_brand_from_text(device_row[field])
                    if extracted_brand:
                        recovered_brand = extracted_brand
                        break
        
        # Try to recover category
        if not recovered_category:
            for field in extraction_fields:
                if field in device_row and not pd.isna(device_row[field]):
                    extracted_category = extract_category_from_text(device_row[field])
                    if extracted_category:
                        recovered_category = extracted_category
                        break
        
        return recovered_brand, recovered_category
    
    # Phase 1: Refine the all_invalid sheet by excluding specific categories
    print("📋 Phase 1: Refining all_invalid sheet by excluding specialized invalid categories...")
    
    # Get the asset tag IDs for exclusion
    unavailable_asset_ids = set(unavailable_inactive_devices['Asset Tag ID'].dropna()) if len(unavailable_inactive_devices) > 0 else set()
    invalid_date_asset_ids = set(invalid_purchase_dates['Asset Tag ID'].dropna()) if len(invalid_purchase_dates) > 0 else set()
    
    # Create refined all_invalid by excluding unavailable and invalid date devices
    if 'all_invalid' in locals() and len(all_invalid) > 0:
        refined_all_invalid = all_invalid[
            ~all_invalid['Asset Tag ID'].isin(unavailable_asset_ids) &
            ~all_invalid['Asset Tag ID'].isin(invalid_date_asset_ids)
        ].copy()
        
        print(f"   Original all_invalid devices: {len(all_invalid)}")
        print(f"   Excluded unavailable/inactive devices: {len(unavailable_asset_ids)}")
        print(f"   Excluded invalid purchase date devices: {len(invalid_date_asset_ids)}")
        print(f"   Refined all_invalid for correction: {len(refined_all_invalid)}")
    else:
        refined_all_invalid = pd.DataFrame()
        print(f"   No all_invalid devices to process")
    
    # Phase 2: Attempt corrections on refined all_invalid
    corrected_devices = []
    correction_stats = {
        'brand_recovered': 0,
        'category_recovered': 0,
        'both_recovered': 0,
        'no_recovery': 0
    }
    
    if len(refined_all_invalid) > 0:
        print(f"\n📋 Phase 2: Attempting data recovery for {len(refined_all_invalid)} devices...")
        
        for idx, device in refined_all_invalid.iterrows():
            original_brand = device['Brand'] if not pd.isna(device['Brand']) and str(device['Brand']).strip() else ""
            original_category = device['Category'] if not pd.isna(device['Category']) and str(device['Category']).strip() else ""
            
            # Attempt recovery
            recovered_brand, recovered_category = attempt_data_recovery(device)
            
            # Update the device record
            device_copy = device.copy()
            brand_recovered = False
            category_recovered = False
            
            if not original_brand and recovered_brand:
                device_copy['Brand'] = recovered_brand
                brand_recovered = True
            
            if not original_category and recovered_category:
                device_copy['Category'] = recovered_category
                category_recovered = True
            
            # Track recovery statistics
            if brand_recovered and category_recovered:
                correction_stats['both_recovered'] += 1
            elif brand_recovered:
                correction_stats['brand_recovered'] += 1
            elif category_recovered:
                correction_stats['category_recovered'] += 1
            else:
                correction_stats['no_recovery'] += 1
            
            # Check if device is now fully valid (has brand AND category)
            final_brand = device_copy['Brand'] if not pd.isna(device_copy['Brand']) and str(device_copy['Brand']).strip() else ""
            final_category = device_copy['Category'] if not pd.isna(device_copy['Category']) and str(device_copy['Category']).strip() else ""
            
            if final_brand and final_category:
                # Device is now corrected - add to corrected list
                corrected_devices.append(device_copy)
        
        print(f"   ✅ Brand recovered: {correction_stats['brand_recovered']} devices")
        print(f"   ✅ Category recovered: {correction_stats['category_recovered']} devices")
        print(f"   ✅ Both recovered: {correction_stats['both_recovered']} devices")
        print(f"   ❌ No recovery possible: {correction_stats['no_recovery']} devices")
        print(f"   🎯 Total devices fully corrected: {len(corrected_devices)}")
    
    # Phase 3: Reclassify corrected data
    if len(corrected_devices) > 0:
        print(f"\n📋 Phase 3: Reclassifying {len(corrected_devices)} corrected devices...")
        
        # Convert corrected devices to DataFrame
        corrected_df = pd.DataFrame(corrected_devices)
        
        # Filter corrected devices to only include those that are also active and have valid purchase dates
        final_corrected = corrected_df[
            (corrected_df['Purchase_Date_Status'] == 'Valid') &
            (corrected_df['Status_Normalized'].str.startswith('ACTIVE'))
        ]
        
        print(f"   💎 Devices ready for fully_valid_data: {len(final_corrected)}")
        
        # Update the fully_valid dataset
        if len(final_corrected) > 0:
            # Remove calculation columns before adding to fully_valid
            final_corrected_clean = final_corrected.drop(columns=['Purchase_Date_Parsed', 'Purchase_Date_Status', 'Device_Age_Years', 'Status_Normalized', 'Issues_Found'], errors='ignore')
            
            # Add corrected devices to enhanced_fully_valid
            enhanced_fully_valid = pd.concat([enhanced_fully_valid, final_corrected_clean], ignore_index=True)
            
            print(f"   🚀 Enhanced fully_valid_data now contains: {len(enhanced_fully_valid)} devices")
            print(f"   � Improvement: +{len(final_corrected_clean)} devices added through correction")
            
            # Update the all_invalid dataset by removing corrected devices
            corrected_asset_ids = set(final_corrected['Asset Tag ID'].dropna())
            if 'all_invalid' in locals():
                remaining_all_invalid = all_invalid[~all_invalid['Asset Tag ID'].isin(corrected_asset_ids)]
                print(f"   🗑️  Remaining uncorrectable invalid devices: {len(remaining_all_invalid)}")
            
        else:
            # No devices were corrected - enhanced_fully_valid already contains the original data
            remaining_all_invalid = all_invalid if 'all_invalid' in locals() else pd.DataFrame()
    else:
        # No corrected devices - enhanced_fully_valid already contains the original data  
        remaining_all_invalid = all_invalid if 'all_invalid' in locals() else pd.DataFrame()
        print(f"\n📋 Phase 3: No devices were corrected - using original fully_valid_data")
    
    # Phase 4: Update final statistics
    print(f"\n📊 === FINAL DATA QUALITY RESULTS (After Advanced Cleaning) ===")
    total_enhanced_valid = len(enhanced_fully_valid)
    original_valid_count = len(original_fully_valid.drop(columns=['Purchase_Date_Parsed', 'Purchase_Date_Status', 'Device_Age_Years', 'Status_Normalized'], errors='ignore'))
    improvement = total_enhanced_valid - original_valid_count
    enhanced_quality_score = round((total_enhanced_valid / len(df)) * 100, 1)
    
    print(f"📊 Original fully valid devices: {original_valid_count}")
    print(f"🎯 Enhanced fully valid devices: {total_enhanced_valid}")
    print(f"📈 Improvement: +{improvement} devices recovered through advanced cleaning")
    print(f"🏆 Enhanced Data Quality Score: {enhanced_quality_score}%")
    
    # Create final analysis-ready count
    final_analysis_ready_devices = enhanced_fully_valid.merge(
        df[['Asset Tag ID', 'Status_Normalized']], 
        on='Asset Tag ID', 
        how='left'
    )
    final_analysis_ready_devices = final_analysis_ready_devices[
        final_analysis_ready_devices['Status_Normalized'].str.startswith('ACTIVE', na=False)
    ]
    
    print(f"✅ Ready for DLM Risk Analysis: {len(final_analysis_ready_devices)} enhanced devices with active status")
    
    # === END ADVANCED CLEANING PROCESS ===
    
    # Save final enhanced results to Excel
    try:
        # Append enhanced results to the existing Excel file
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
            # Enhanced fully valid data
            if len(enhanced_fully_valid) > 0:
                enhanced_fully_valid.to_excel(writer, sheet_name='Enhanced_Fully_Valid_Data', index=False)
            
            # Remaining invalid data after correction attempts
            if len(remaining_all_invalid) > 0:
                remaining_all_invalid.to_excel(writer, sheet_name='Remaining_Invalid_Data', index=False)
        
        # Apply color formatting to the new sheets
        print("🎨 Applying color formatting to new enhanced sheets...")
        
        try:
            workbook = openpyxl.load_workbook(output_path)
            
            # Color formatting for the new sheets
            new_sheet_colors = {
                'Enhanced_Fully_Valid_Data': ('27AE60', 'D5F4E6'),  # Bright green - enhanced valid data
                'Remaining_Invalid_Data': ('C0392B', 'F5B7B1'),     # Bright red - remaining invalid data
            }
            
            # Apply formatting to new sheets
            for sheet_name, (header_color, data_color) in new_sheet_colors.items():
                if sheet_name in workbook.sheetnames:
                    success = apply_sheet_formatting(workbook, sheet_name, header_color, data_color)
                    if success:
                        print(f"  ✅ Formatted {sheet_name}")
                    else:
                        print(f"  ❌ Failed to format {sheet_name}")
                else:
                    print(f"  ⚠️  Sheet {sheet_name} not found")
            
            # Save the formatted workbook
            workbook.save(output_path)
            workbook.close()
            print(f"🎨 Applied color formatting to enhanced sheets!")
            
        except Exception as e:
            print(f"❌ Error during enhanced sheet color formatting: {e}")
        
        print(f"📂 Final results saved to {output_path}")
        
    except PermissionError:
        print(f"\nERROR: Permission denied when trying to save to {output_path}")
        print("This usually means:")
        print("1. The file is currently open in Excel or another application - please close it")
        print("2. You don't have write permissions to the folder")
        print("3. The file is locked by another process")
        print("\nPlease close any open Excel files and try again.")
        
    except Exception as e:
        print(f"\nError saving final results: {e}")

if __name__ == "__main__":
    main()
