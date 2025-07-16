import pandas as pd
import datetime
from dateutil import parser
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

# -------------------------------
# 1. Load and Clean the CSV Data
# -------------------------------
def load_csv(filepath):
    return pd.read_csv(filepath)

def infer_device_type(row):
    text = f"{row.get('model', '')} {row.get('description', '')} {row.get('brand', '')}".lower()
    if 'laptop' in text or 'notebook' in text:
        return 'Laptop'
    elif 'ipad' in text or 'tablet' in text:
        return 'Tablet'
    elif 'monitor' in text or 'display' in text:
        return 'Monitor'
    elif 'dock' in text:
        return 'Docking Station'
    elif 'phone' in text:
        return 'Phone'
    elif 'access point' in text or 'ap' in text:
        return 'Access Point'
    elif 'desktop' in text or 'prodesk' in text:
        return 'Desktop'
    elif 'voip' in text:
        return 'VoIP Phone'
    else:
        return 'Other'

def infer_os(row):
    text = f"{row.get('model', '')} {row.get('description', '')}".lower()
    if 'mac' in text or 'apple' in text:
        return 'mac'
    elif 'win11' in text or 'windows 11' in text:
        return 'windows_11'
    elif 'win10' in text or 'windows 10' in text:
        return 'windows_10'
    elif 'chrome' in text or 'chromebook' in text:
        return 'chrome_os'
    else:
        return 'unknown'

def clean_data(df, log_file='cleaning_log.txt'):
    df.columns = [col.strip().lower().replace(' ', '_') for col in df.columns]

    with open(log_file, 'w', encoding='utf-8') as log:
        log.write("[Cleaned columns:]\n")
        log.write(f"{df.columns.tolist()}\n\n")

        required_cols = ['asset_tag_id', 'purchase_date', 'brand', 'created_by']
        existing_cols = [col for col in required_cols if col in df.columns]

        if len(existing_cols) < len(required_cols):
            missing = list(set(required_cols) - set(existing_cols))
            log.write(f"[Missing required columns:] {missing}\n")

        if existing_cols:
            df = df.dropna(subset=existing_cols)
            log.write(f"[Dropped rows with missing values in:] {existing_cols}\n")
        else:
            log.write("[No required columns found. Skipping dropna.]\n")

    df['device_type'] = df.apply(infer_device_type, axis=1)
    df['os'] = df.apply(infer_os, axis=1)
    return df

# --------------------------------------
# 2. Risk Scoring Criteria and Functions
# --------------------------------------
def categorize_device_age(purchase_date):
    age = (datetime.datetime.now() - parser.parse(purchase_date)).days / 365
    if age >= 5:
        return 'High Risk', 3, 'Age >5 yrs (+3)'
    elif 3 <= age < 5:
        return 'Medium Risk', 2, 'Age 3-5 yrs (+2)'
    else:
        return 'Low Risk', 1, 'Age <3 yrs (+1)'

def score_os_lifecycle(os_name):
    return {
        'windows_11': (1, ''),
        'windows_10': (2, ''),
        'windows_7': (3, ''),
        'mac': (2, ''),
        'chrome_os': (2, ''),
        'unknown': (3, '')
    }.get(os_name.lower(), (3, ''))

def score_ownership(owner_field):
    return 2 if 'shared' in str(owner_field).lower() else 1

def score_brand(brand):
    return {'apple': 1, 'microsoft': 2, 'hp': 2, 'lenovo': 2, 'lg': 3}.get(str(brand).lower(), 2)

def score_device_type(device_type):
    reasoning = ''
    if device_type.lower() in ['laptop', 'phone', 'tablet']:
        return 1, 'High-turnover category (+1)'
    return 0, ''

# ------------------------------------------------
# 3. Total Score Calculation and Risk Categorizing
# ------------------------------------------------
def calculate_total_risk(row):
    try:
        _, age_score, age_reason = categorize_device_age(row['purchase_date'])
        os_score, _ = score_os_lifecycle(row['os'])
        brand_score = score_brand(row['brand'])
        type_score, type_reason = score_device_type(row['device_type'])

        total_score = age_score + brand_score + type_score

        if total_score >= 5:
            risk_level = 'High Risk'
        elif total_score >= 3:
            risk_level = 'Medium Risk'
        else:
            risk_level = 'Low Risk'

        warranty = 36 if row['device_type'].lower() in ['laptop', 'desktop'] else 12

        reasons = [age_reason]
        if warranty == 12:
            reasons.append('Warranty expired (+2)')
        if type_reason:
            reasons.append(type_reason)

        return pd.Series({
            'Asset Tag': row.get('asset_tag_id', 'Unknown'),
            'Asset Name': row.get('model', 'Unknown'),
            'Category': row['device_type'],
            'Purchase Date': row['purchase_date'],
            'Warranty': warranty,
            'Risk Score': total_score,
            'Risk Level': risk_level,
            'Reasoning': ', '.join(reasons)
        })

    except Exception as e:
        return pd.Series({
            'Asset Tag': 'Error',
            'Asset Name': 'Error',
            'Category': '',
            'Purchase Date': '',
            'Warranty': '',
            'Risk Score': '',
            'Risk Level': 'Error',
            'Reasoning': str(e)
        })

# --------------------------
# 4. Apply and Export Result
# --------------------------
def apply_risk_analysis(df):
    results_df = df.apply(calculate_total_risk, axis=1)
    return results_df

def export_results(df, filename='processed_inventory.xlsx'):
    selected_cols = ['Asset Tag', 'Asset Name', 'Category', 'Purchase Date', 'Warranty', 'Risk Score', 'Risk Level', 'Reasoning']

    wb = Workbook()
    del wb[wb.sheetnames[0]]  # Remove default sheet

    fill_colors = {
        'High Risk': 'FFC7CE',    # Light red
        'Medium Risk': 'FFEB9C',  # Light yellow
        'Low Risk': 'C6EFCE'      # Light green
    }

    for risk_level in ['High Risk', 'Medium Risk', 'Low Risk']:
        risk_df = df[df['Risk Level'] == risk_level][selected_cols]
        ws = wb.create_sheet(title=risk_level)

        for r in dataframe_to_rows(risk_df, index=False, header=True):
            ws.append(r)

        for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.fill = PatternFill(start_color=fill_colors[risk_level], end_color=fill_colors[risk_level], fill_type="solid")

    wb.save(filename)

# --------------
# 5. Main Script
# --------------
def main():
    df = load_csv('assets.csv')
    df = clean_data(df, log_file='data_cleaning_output.txt')
    df_result = apply_risk_analysis(df)
    export_results(df_result)
    print("Risk analysis complete. Results saved to 'processed_inventory.xlsx'.")
    print("Data cleaning log saved to 'data_cleaning_output.txt'.")

if __name__ == "__main__":
    main()
